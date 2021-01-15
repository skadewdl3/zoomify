import openpyxl
import time
import pyautogui as gui
import os
import datetime
import schedule
import pyperclip
import json
from multiprocessing import Process

CONFIDENCE = 0.9
DATA_PATH = os.path.abspath('./Meeting.xlsx')
book = openpyxl.load_workbook(DATA_PATH)
sheet = book.active
ROWS = sheet.max_row + 1
COLUMNS = sheet.max_column
past_meetings = []
meeting_data = []
future_meetings = []
meetings_json = {}

def open_zoom():
  gui.press('win', interval=0.5)
  gui.write('zoom')
  time.sleep(2)
  gui.press('enter', interval=0.5)

def sign_in():
  res = gui.locateCenterOnScreen('images/sign-in.png')
  if res == None:
    print('returning')
    return False
  x, y = gui.locateCenterOnScreen('images/join-4.png')
  gui.click(x, y)
  return True

def update_excel():
  sheet.delete_rows(2, sheet.max_row - 1)
  res = meetings_json
  access_row = 2
  arr = ['past', 'current', 'future']
  for type in arr:
    for meeting in res[type]:
      sheet.cell(row = access_row, column = 1).value = meeting['date']
      sheet.cell(row = access_row, column = 2).value = f'{meeting["time"]}'
      sheet.cell(row = access_row, column = 3).value = meeting['id']
      sheet.cell(row = access_row, column = 4).value = meeting['password']
      sheet.cell(row = access_row, column = 5).value = meeting['link']
      sheet.cell(row = access_row, column = 6).value = meeting['duration']
      access_row += 1
  book.save('Meeting.xlsx')

def get_json():
  update_json()
  print(meetings_json)
  return meetings_json

def delete_meeting(type, index):
  global meetings_json
  res = get_json()
  res[type].pop(index)
  print(type, index)
  meetings_json = res
  update_excel()
  return res

def create_meeting(data):
  print('creating meeting')

def update_json():
  global meetings_json
  def save_meeting_to_arr(meeting, arr):
    arr.append({
      'date': meeting['date'].strftime('%d-%m-%Y'),
      'time': meeting['time'].strftime('%H:%M:%S'),
      'id': meeting['id'],
      'password': meeting['password'],
      'link': meeting['link'],
      'duration': meeting['duration']
    })

  meetings_arr = []
  past_meetings_arr = []
  future_meetings_arr = []


  for meeting in meeting_data:
    save_meeting_to_arr(meeting, meetings_arr)
  
  for meeting in past_meetings:
    save_meeting_to_arr(meeting, past_meetings_arr)
  
  for meeting in future_meetings:
    save_meeting_to_arr(meeting, future_meetings_arr)
  
  meetings_json['current'] = meetings_arr
  meetings_json['past'] = past_meetings_arr
  meetings_json['future'] = future_meetings_arr

def parse_meeting_data():
  global past_meetings 
  global meeting_data 
  global future_meetings 
  global meetings_json 
  for row in range(ROWS):    
    if row == 0 or row == 1:
      past_meetings = []
      meeting_data = []
      future_meetings = []
      meetings_json = {}
      continue
    date_cell = sheet.cell(row = row, column = 1)
    time_cell = sheet.cell(row = row, column = 2)
    id_cell = sheet.cell(row = row, column = 3)
    password_cell = sheet.cell(row = row, column = 4)
    link_cell = sheet.cell(row = row, column = 5)
    duration_cell = sheet.cell(row = row, column = 6)
    current_date = datetime.datetime.now()
    current_day = current_date.strftime('%d')
    current_month = current_date.strftime('%m')
    current_year = current_date.strftime('%Y')
    print(date_cell.value)
    day = date_cell.value.strftime('%d')
    month = date_cell.value.strftime('%m')
    year = date_cell.value.strftime('%Y')
    
    meeting = {
      'date': date_cell.value,
      'time': time_cell.value,
      'id': id_cell.value,
      'password': password_cell.value,
      'link': link_cell.value,
      'duration': duration_cell.value
    }
    if current_day == day and current_month == month and current_year == year:
      meeting_data.append(meeting)
    elif current_day > day or current_month > month or current_year > year:
      past_meetings.append(meeting)
    elif current_day < day or current_month < month or current_year < year:
      future_meetings.append(meeting)

def join_meeting(meeting):
  link = meeting['link']
  password = meeting['password']
  id = meeting['id']
  open_zoom()
  time.sleep(2)
  direct_login = sign_in()
  if direct_login == False:
    time.sleep(10)
    res = gui.locateCenterOnScreen('images/join.png', confidence = CONFIDENCE)
    x, y = res
    gui.click(x, y)
    time.sleep(2)
  else:
    time.sleep(12)
  if link == None:
    gui.typewrite(str(id), interval=0.1)
    gui.press('tab', interval=0.5) 
    gui.press('tab', interval=0.5) 
  else:
    pyperclip.copy(link)
    gui.hotkey('ctrl', 'v')
    gui.press('tab', interval=0.5) 
    gui.press('tab', interval=0.5)
  gui.hotkey('ctrl', 'a')
  gui.press('backspace')
  gui.typewrite('819 - Soham Karandikar', interval=0.1)
  x, y= gui.locateCenterOnScreen('images/turn-off-video.png', confidence = CONFIDENCE)
  gui.click(x, y)
  x, y= gui.locateCenterOnScreen('images/join-meeting.png', confidence = CONFIDENCE)
  gui.click(x, y)
  if password != None:
    time.sleep(5)
    gui.typewrite(password, interval=0.1)
    x, y = gui.locateCenterOnScreen('images/join-3.png', confidence = CONFIDENCE)
    gui.click(x, y)
  time.sleep(10)
  while True:
    res = gui.locateCenterOnScreen('images/close.png', confidence = CONFIDENCE)
    if res == None:
      break
    else:
      time.sleep(5)

def start_screen_recorder():
  while True:
    res = gui.locateCenterOnScreen('images/unmute.png', confidence = CONFIDENCE)
    if res != None:
      break
    else:
      time.sleep(5)
  gui.hotkey('alt', 'f')
  gui.press('win')
  gui.write('obs', interval=0.1)
  gui.press('enter')
  time.sleep(15)
  x, y = gui.locateCenterOnScreen('./images/record.png', confidence = CONFIDENCE)
  gui.click(x, y)
  gui.hotkey('alt', 'tab')
  
def stop_screen_recorder():
  gui.hotkey('alt', 'tab')
  time.sleep(5)
  x, y = gui.locateCenterOnScreen('./images/stop-record.png', confidence = CONFIDENCE)
  gui.click(x, y)
  time.sleep(2)
  gui.hotkey('alt', 'f4')

def exit_meeting():
  gui.hotkey('alt', 'q')
  gui.press('tab')
  gui.press('enter')

def schedule_classes():
  print(meeting_data)
  for meeting in meeting_data:
    def start_meeting_and_record(): 
      join_meeting(meeting)
      start_screen_recorder()
      time.sleep(int(meeting['duration']) * 60)
      stop_screen_recorder()
      time.sleep(2)
      exit_meeting()
    print(meeting['time'].strftime('%H:%M'))
    schedule.every().day.at(meeting['time'].strftime('%H:%M')).do(start_meeting_and_record)
  
def run_scheduler():
  while True:
    print('running')
    schedule.run_pending()
    time.sleep(5)

def init(email, password):
  global book
  global sheet
  global ROWS
  global COLUMNS
  book = openpyxl.load_workbook(DATA_PATH)
  sheet = book.active
  ROWS = sheet.max_row + 1
  COLUMNS = sheet.max_column
  print(email, password)
  parse_meeting_data()
  schedule_classes()

